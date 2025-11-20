#!/usr/bin/env python3
"""
Excel export functionality for RateMySite analysis
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_report(results: List[Dict[str, Any]], filepath: str, table_rows: List[Tuple[str, str]]) -> None:
    """
    Create a formatted Excel report from analysis results
    
    Args:
        results: List of dictionaries containing analysis data
        filepath: Path where Excel file should be saved
        table_rows: List of (key, display_name) tuples defining the structure
    """
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "RateMySite Analysis"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    score_font = Font(bold=True)
    
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    # Add title and metadata
    ws['A1'] = "RateMySite Analysis Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Total sites analyzed: {len(results)}"
    
    # Start data table at row 5
    start_row = 5
    current_row = start_row
    
    if not results:
        ws[f'A{current_row}'] = "No results to display"
        wb.save(filepath)
        return
    
    # Create headers - Category column + one column per analyzed site
    ws[f'A{current_row}'] = "Category"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].alignment = center_alignment
    ws[f'A{current_row}'].border = border
    
    # Add URL headers
    for col_idx, result in enumerate(results, start=2):
        col_letter = chr(ord('A') + col_idx - 1)  # B, C, D, etc.
        url = result.get('URL', 'Unknown')
        
        # Try to extract domain name for cleaner headers
        try:
            if url.startswith(('http://', 'https://')):
                domain = url.split('//')[1].split('/')[0]
                # Remove www. if present
                domain = domain.replace('www.', '')
            else:
                domain = url
        except:
            domain = url
            
        ws[f'{col_letter}{current_row}'] = domain
        ws[f'{col_letter}{current_row}'].font = header_font
        ws[f'{col_letter}{current_row}'].fill = header_fill
        ws[f'{col_letter}{current_row}'].alignment = center_alignment
        ws[f'{col_letter}{current_row}'].border = border
    
    current_row += 1
    
    # Add data rows based on table_rows structure
    for row_key, row_display in table_rows:
        # Category name in column A
        ws[f'A{current_row}'] = row_display
        ws[f'A{current_row}'].font = subheader_font
        ws[f'A{current_row}'].fill = subheader_fill
        ws[f'A{current_row}'].alignment = left_alignment
        ws[f'A{current_row}'].border = border
        
        # Data for each site
        for col_idx, result in enumerate(results, start=2):
            col_letter = chr(ord('A') + col_idx - 1)
            
            value = result.get(row_key, '-')
            if value is None:
                value = '-'
            
            cell = ws[f'{col_letter}{current_row}']
            cell.value = str(value)
            cell.border = border
            cell.alignment = center_alignment
            
            # Special formatting for scores
            if 'Score' in row_key and value != '-' and str(value).isdigit():
                cell.font = score_font
                score = int(value)
                if score >= 80:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
                elif score >= 60:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
                elif score < 60:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
            
            # Special formatting for URLs (make them clickable)
            elif row_key == 'URL' and value != '-':
                cell.hyperlink = value
                cell.font = Font(color="0000FF", underline="single")
        
        current_row += 1
    
    # Add summary section
    current_row += 2
    ws[f'A{current_row}'] = "Summary Statistics"
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    current_row += 1
    
    # Calculate average scores
    score_fields = [key for key, _ in table_rows if 'Score' in key]
    
    for score_field in score_fields:
        scores = []
        for result in results:
            value = result.get(score_field, '-')
            if value != '-' and str(value).isdigit():
                scores.append(int(value))
        
        if scores:
            avg_score = sum(scores) / len(scores)
            ws[f'A{current_row}'] = f"Average {score_field}:"
            ws[f'B{current_row}'] = f"{avg_score:.1f}"
            current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set minimum and maximum widths
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(filepath)

def create_detailed_excel_report(results: List[Dict[str, Any]], filepath: str) -> None:
    """
    Create a detailed Excel report with multiple sheets
    
    Args:
        results: List of dictionaries containing analysis data
        filepath: Path where Excel file should be saved
    """
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        
        # Summary sheet
        summary_data = []
        for result in results:
            row = {
                'URL': result.get('URL', ''),
                'Company': result.get('Company', ''),
                'Overall Score': result.get('Overall Score', ''),
                'Consumer Score': result.get('Consumer Score', ''),
                'Developer Score': result.get('Developer Score', ''),
                'Investor Score': result.get('Investor Score', ''),
                'Trust Score': result.get('Trust Score', ''),
                'UX Score': result.get('UX Score', ''),
            }
            summary_data.append(row)
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Detailed data sheet
        if results:
            detailed_df = pd.DataFrame(results)
            # Remove raw data column if it exists (too verbose for Excel)
            if '_raw' in detailed_df.columns:
                detailed_df = detailed_df.drop('_raw', axis=1)
            detailed_df.to_excel(writer, sheet_name='Detailed Data', index=False)
        
        # Scores comparison sheet (only numeric scores)
        scores_data = []
        score_columns = ['Overall Score', 'Consumer Score', 'Developer Score', 
                        'Investor Score', 'Clarity Score', 'Visual Design Score', 
                        'UX Score', 'Trust Score', 'Value Prop Score']
        
        for result in results:
            row = {'URL': result.get('URL', '')}
            for col in score_columns:
                value = result.get(col, '')
                # Convert to numeric if possible
                try:
                    row[col] = int(value) if value and str(value).isdigit() else None
                except:
                    row[col] = None
            scores_data.append(row)
        
        if scores_data:
            scores_df = pd.DataFrame(scores_data)
            scores_df.to_excel(writer, sheet_name='Scores Comparison', index=False)
